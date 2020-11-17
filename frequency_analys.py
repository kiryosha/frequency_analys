import re
import os
import threading
import subprocess
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def writeLog(e):
	log = open('./log.txt','a');
	log.write(str(e) + '\n');
	log.close();

def searchDisks():
	fsinfo = subprocess.check_output('fsutil fsinfo drives');
	disks = re.findall("[A-Z]", fsinfo.decode('ISO-8859-1'))
	return disks;

def writingFile(disk):
	size = 0;
	count = 0;
	diskName = './files' + disk + '.txt';
	file = open(diskName, 'w', encoding='utf-8');
	for root, dirs, files in os.walk(disk + ':/'):
		for name in files:
			file.write(os.path.join(root, name) + '\n');
			try:
				size += os.path.getsize(os.path.join(root, name));
			except Exception as e:
				writeLog(e);
			count += 1;
	file.close();
	return (diskName, count, size);

def initDict():
	slovar = {};
	for symbolLeft in range(256):
		for symbolRigth in range(256):
			slovar.update({(symbolLeft,symbolRigth): 0});
	return slovar;

def searchBytes(files, countDisks, fileOut, sizeFull):
	count = 1;
	sizeNow = 0;
	slovar = initDict();
	file = files;
	files = open(file, 'r', encoding='utf-8');
	for line in files:
		try:
			data = open(line[:-1], 'rb');
			text = data.read();
			for i in range(len(text)-1):
				if((text[i], text[i+1]) in slovar):
					slovar[(text[i], text[i+1])] +=1;
			try:
				sizeNow += os.path.getsize(line[:-1]);
			except Exception as e:
				writeLog(e);
			print('[' + str(count) + ';' + str(countDisks) +'] ' + str(round(sizeNow * 0.000000001, 3)) + '/' + str(round(sizeFull * 0.000000001, 3)) + 'Gb Disk: ' + line[0]);
		except Exception as e:
			try:
				writeLog(e);
				sizeFull -= os.path.getsize(line[:-1]);
			except Exception as ex:
				print(ex);
		count += 1;
	with open(fileOut,'w') as out:
		for key,val in slovar.items():
			out.write('{} : {}\n'.format(key,val));
	out.close();


def convertTuple(tup): 
    result =  ''.join(str(tup)); 
    return result;

def printResult(resultFiles):
	bytesArr = [];
	for symbolLeft in range(256):
		for symbolRigth in range(256):
			bytesArr.append((symbolLeft,symbolRigth));
	countArr = [0 for i in range(256*256)];
	for file in resultFiles:
		data = open(file);
		i = 0;
		for line in data:
			temp = re.findall(r'\d+$', line);
			countArr[i] += int(temp[0]);
			i += 1;
		data.close();
	file = open('./result.txt', 'w');
	i = 0;
	for temp in bytesArr:
		count = str(countArr[i]);
		file.write(convertTuple(temp) + " - " + count + '\n');
		i += 1;
	file.close();

def wrireToXlsx():
	countLine = 0;
	countArr = [];
	bytesArr = []
	wb = Workbook();
	ws = wb.active;
	for i in range(1,3):
		ws.column_dimensions[get_column_letter(i)].width = 12;
	ws.title = 'Result';
	file = open('./result.txt', 'r');
	for line in file:
		countLine += 1;
		countArr.append(re.findall(r'\d+$', line)[0]);
		bytesArr.append('(' + re.findall(r'(\d*, \d*)', line)[0] + ')');
	for i in range(1, countLine + 1):
		ws.cell(row= i, column= 1, value = bytesArr[i-1]);
		ws.cell(row= i, column= 2, value = int(countArr[i-1]));
	file.close();
	wb.save('result.xlsx')

def delRubbish(file, fileOutNameArr):
	for i in range(len(file)):
		path = os.path.join(os.path.abspath(os.path.dirname(__file__)), file[i]);
		os.remove(path);
		path = os.path.join(os.path.abspath(os.path.dirname(__file__)), fileOutNameArr[i]);
		os.remove(path);

def main():
	size = [];
	file = [];
	diskCount = [];
	fileOutNameArr = [];
	disksList = searchDisks();
	for name in disksList:
		tmp = writingFile(name);
		diskCount.append(tmp[1]);
		file.append(tmp[0]);
		size.append(tmp[2]);
		fileOut = './out' + name + '.txt';
		fileOutNameArr.append(fileOut);
	thd = [i for i in range(len(disksList))]
	for i in range(len(disksList)):
		thd[i] = threading.Thread(target=searchBytes, args=(file[i], diskCount[i], fileOutNameArr[i], size[i]));
		thd[i].start();
	for i in range(len(disksList)):
		thd[i].join();
	printResult(fileOutNameArr);
	wrireToXlsx();
	delRubbish(file, fileOutNameArr);


if __name__ == "__main__":
	main()

